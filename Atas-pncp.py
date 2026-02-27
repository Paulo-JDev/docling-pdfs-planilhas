import os
import re
import asyncio
import aiohttp
import openpyxl
import pandas as pd
from pathlib import Path
from tqdm.asyncio import tqdm_asyncio
from tqdm import tqdm

# --- CONFIGURAÇÕES ---
INPUT_PLANILHA = "Relatorio_Atas_Administrativas-2.xlsx" 
PASTA_DOWNLOADS = "atas_baixadas"
PASTA_ERROS = "erros_md"
ARQUIVO_SAIDA = "resultado_atas.csv"

# --- REGEX ---
REGEX_VALOR = re.compile(r'(?:Valor\s+total.*?|Total.*?|R\$)\s*R?\$?\s*(\d{1,3}(?:\.\d{3})*(?:,\d{1,2})?)', re.IGNORECASE | re.DOTALL)
REGEX_RENOVACAO_QTD = re.compile(r'5\.1\.1.*?prorrogação.*?renovado.*?quantitativo', re.IGNORECASE | re.DOTALL)

def setup_folders():
    Path(PASTA_DOWNLOADS).mkdir(exist_ok=True)
    Path(PASTA_ERROS).mkdir(exist_ok=True)

# ---------------------------------------------------------
# 1. EXTRAÇÃO DE LINKS DA PLANILHA
# ---------------------------------------------------------
def extrair_links_da_planilha(caminho_planilha):
    print(f"🔍 Lendo planilha: {caminho_planilha}...")
    
    if not os.path.exists(caminho_planilha):
        print(f"❌ ERRO: O arquivo '{caminho_planilha}' não foi encontrado na pasta atual.")
        return []

    try:
        wb = openpyxl.load_workbook(caminho_planilha, data_only=False)
        ws = wb.active
    except Exception as e:
        print(f"❌ Erro ao abrir a planilha com openpyxl: {e}")
        return []

    links_encontrados = []
    
    print("   Processando linhas do Excel...")
    for row in range(2, ws.max_row + 1):
        celula_ata = ws.cell(row=row, column=6)
        texto_ata = str(celula_ata.value).strip() if celula_ata.value else ""
        url = None
        
        if celula_ata.hyperlink:
            url = celula_ata.hyperlink.target
            
        if url and url.startswith("http"):
            nome_limpo = texto_ata.replace('/', '-').replace('\\', '-').replace(' ', '_')
            if not nome_limpo or nome_limpo == "None":
                nome_limpo = f"ata_linha_{row}"
                
            links_encontrados.append({
                "nome_ata": nome_limpo,
                "url_pncp": url
            })
            
    print(f"✅ Encontrados {len(links_encontrados)} links válidos.")
    return links_encontrados

# ---------------------------------------------------------
# 2. DOWNLOADER COM SISTEMA ANTI-BLOQUEIO
# ---------------------------------------------------------
async def baixar_pdf_pncp(session, item, sem):
    # O Semaphore funciona como uma catraca: só deixa X passarem por vez
    async with sem:
        url_original = item['url_pncp']
        nome_base = item['nome_ata']
        caminho_final = Path(PASTA_DOWNLOADS) / f"{nome_base}.pdf"

        if caminho_final.exists():
            return str(caminho_final)

        # MÁGICA: Transforma o link da tela no link direto da API
        match_ata = re.search(r'pncp\.gov\.br/app/atas/(\d+)/(\d+)/(\d+)/(\d+)', url_original)
        if match_ata:
            cnpj, ano, compra, ata = match_ata.groups()
            link_download = f"https://pncp.gov.br/pncp-api/v1/orgaos/{cnpj}/compras/{ano}/{compra}/atas/{ata}/arquivos/1"
        else:
            match_edital = re.search(r'pncp\.gov\.br/app/editais/(\d+)/(\d+)/(\d+)', url_original)
            if match_edital:
                cnpj, ano, compra = match_edital.groups()
                link_download = f"https://pncp.gov.br/pncp-api/v1/orgaos/{cnpj}/compras/{ano}/{compra}/arquivos/1"
            else:
                link_download = url_original

        try:
            # Pausa de 0.5 segundos para não irritar o firewall do Governo
            await asyncio.sleep(0.5) 
            
            async with session.get(link_download, timeout=30) as resp_pdf:
                if resp_pdf.status == 200:
                    content = await resp_pdf.read()
                    if len(content) > 1000:
                        with open(caminho_final, 'wb') as f:
                            f.write(content)
                        return str(caminho_final)
            return None
                
        except Exception:
            return None

async def gerenciar_downloads(lista_links):
    # Catraca: Libera no máximo 3 downloads simultâneos
    sem = asyncio.Semaphore(3) 
    
    timeout = aiohttp.ClientTimeout(total=60)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        tasks = [baixar_pdf_pncp(session, item, sem) for item in lista_links]
        # O erro estava aqui. Removido o return_exceptions=True!
        resultados = await tqdm_asyncio.gather(*tasks, desc="⬇️ Baixando Atas")
        
    caminhos = []
    for res in resultados:
        if isinstance(res, str):
            caminhos.append(res)
            
    return caminhos

# ---------------------------------------------------------
# 3. ANÁLISE DO PDF COM DOCLING
# ---------------------------------------------------------
def converter_str_valor(valor_str):
    try:
        limpo = valor_str.replace('.', '').replace(',', '.')
        return float(limpo)
    except:
        return 0.0

def analisar_pdf_individual(pdf_path):
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        return {"Arquivo": Path(pdf_path).name, "Status": "Erro: Biblioteca docling não instalada."}

    nome_arquivo = Path(pdf_path).name
    resultado = {
        "Arquivo": nome_arquivo,
        "Valor Total": 0.0,
        "Renovacao Vigencia": "S",
        "Renovacao Quantitativo (5.1.1)": "NÃO",
        "Status": "Sucesso"
    }

    try:
        converter = DocumentConverter()
        doc = converter.convert(pdf_path)
        markdown_text = doc.document.export_to_markdown()

        matches_valor = REGEX_VALOR.findall(markdown_text)
        valor_encontrado = False
        max_val = 0
        
        if matches_valor:
            valores_float = [converter_str_valor(v) for v in matches_valor]
            max_val = max(valores_float) if valores_float else 0
            resultado["Valor Total"] = f"R$ {max_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            valor_encontrado = True
        
        if REGEX_RENOVACAO_QTD.search(markdown_text):
            resultado["Renovacao Quantitativo (5.1.1)"] = "SIM"
        
        if not valor_encontrado or max_val == 0:
            resultado["Status"] = "Verificar Valor"
            md_path = Path(PASTA_ERROS) / f"{nome_arquivo}.md"
            with open(md_path, "w", encoding="utf-8") as f:
                f.write(markdown_text)

    except Exception:
        resultado["Status"] = "Erro de leitura no PDF"
    
    return resultado

# ---------------------------------------------------------
# 4. ORQUESTRAÇÃO PRINCIPAL
# ---------------------------------------------------------
def main():
    setup_folders()
    versao = 1.0 
    print(f"Versão do Programa: {versao}\n")
    print("--- INICIANDO ROBÔ DE ATAS ---")
    
    links = extrair_links_da_planilha(INPUT_PLANILHA)
    if not links:
        return

    arquivos_pdf = asyncio.run(gerenciar_downloads(links))
    
    if not arquivos_pdf:
        print("\n❌ Nenhum PDF foi baixado com sucesso.")
        return

    print("\n" + "="*60)
    print("⏳ INICIANDO A INTELIGÊNCIA ARTIFICIAL (DOCLING)")
    print("⚠️  Atenção: O computador pode parecer travado por alguns minutos.")
    print("    NÃO APERTE CTRL+C, APENAS AGUARDE!")
    print("="*60 + "\n")
    
    resultados_finais = []
    for pdf in tqdm(arquivos_pdf, desc="⚙️ Analisando PDFs"):
        res = analisar_pdf_individual(pdf)
        resultados_finais.append(res)

    df = pd.DataFrame(resultados_finais)
    df.to_csv(ARQUIVO_SAIDA, index=False, sep=';', encoding='utf-8-sig')
    
    print("\n" + "="*40)
    print("✅ PROCESSAMENTO CONCLUÍDO COM SUCESSO!")
    print("="*40)
    print(f"📄 Relatório salvo em: {ARQUIVO_SAIDA}")
    print(f"⚠️  PDFs que precisam de auditoria humana salvos em: {PASTA_ERROS}/")

if __name__ == "__main__":
    main()
